# tutorial: https://openpyxl.readthedocs.io/en/default/tutorial.html#accessing-many-cells

# pandas read_excel
df = pd.read_excel('test.xlsx')


# Using openpyxl
import openpyxl

# read excel
wb = openpyxl.load_workbook(my_excel)  # read excel file
my_sheet = wb.get_sheet_by_name(sheet_name) # read sheet
col = my_sheet['A4':'A10']  # read multiple rows in a col, here 'A', 'B'.. indicates cols in Excel
for elem in col:
    print elem[0].value # access to the cell value

    
# write excel, with multiple tabs
## The benefit here is, if there are columns in JSON format or has so many characters in it, that using whatever delimiter
### to export from SQL Workbench won't make the excel sheets reader friendly. Output through python with this code can solve 
### all the problems.
import openpyxl 
import os

def write_excel(df, tab_idx, tab_title, output_file):
    
    if os.path.isfile(output_file):
        wb = openpyxl.load_workbook(output_file)
        ws = wb.create_sheet(tab_title)
    else:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(tab_title)
    
    # write header first
    header = df.columns
    for i in range(len(header)):
        c = ws.cell(row=1, column=i+1)
        c.value = header[i]
        
    # write values
    for i in range(df.shape[0]):
        for j in range(len(header)):
            c = ws.cell(row=2+i, column=j+1)
            c.value = df.iloc[i][j]
    
    wb.save(output_file)
    
write_excel(df1, 1, 'tab1', 'my_output.xlsx')
write_excel(df2, 2, 'tab2', 'my_output.xlsx')
    
